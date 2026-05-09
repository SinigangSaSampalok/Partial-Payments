from __future__ import annotations

import calendar
import re
from datetime import datetime
from pathlib import Path
import shutil
import tkinter as tk
import ttkbootstrap as tb
from ttkbootstrap.constants import BOTH, EW, LEFT, RIGHT, W, X, Y
from tkinter import filedialog, messagebox


def _is_valid_numeric_input(value: str) -> bool:
    if value == "":
        return True
    return re.fullmatch(r"\d*(\.\d{0,2})?", value) is not None


def _center_dialog(window: tb.Toplevel, parent) -> None:
    window.update_idletasks()

    width = window.winfo_width()
    height = window.winfo_height()
    req_width = window.winfo_reqwidth()
    req_height = window.winfo_reqheight()
    width = width if width > 1 else req_width
    height = height if height > 1 else req_height

    if parent is not None and parent.winfo_exists() and parent.winfo_ismapped():
        parent.update_idletasks()
        parent_width = parent.winfo_width()
        parent_height = parent.winfo_height()

        if parent_width > 1 and parent_height > 1:
            x_pos = parent.winfo_rootx() + (parent_width - width) // 2
            y_pos = parent.winfo_rooty() + (parent_height - height) // 2
        else:
            x_pos = (window.winfo_screenwidth() - width) // 2
            y_pos = (window.winfo_screenheight() - height) // 2
    else:
        x_pos = (window.winfo_screenwidth() - width) // 2
        y_pos = (window.winfo_screenheight() - height) // 2

    window.geometry(f"{width}x{height}+{max(0, x_pos)}+{max(0, y_pos)}")


def _center_when_ready(window: tb.Toplevel, parent) -> None:
    window.after(10, lambda: _center_dialog(window, parent))
    window.after(120, lambda: _center_dialog(window, parent))


class ClientFormDialog(tb.Toplevel):
    def __init__(
        self,
        parent,
        title: str,
        on_submit,
        client: dict | None = None,
        show_item: bool = True,
        show_balance: bool = True,
    ) -> None:
        super().__init__(parent)
        self.title(title)
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()

        self._on_submit = on_submit
        self._show_item = show_item
        self._show_balance = show_balance
        self.client_id = client.get("id") if client else None
        self._numeric_vcmd = (self.register(_is_valid_numeric_input), "%P")

        container = tb.Frame(self, padding=18)
        container.pack(fill=BOTH, expand=True)

        row = 0
        tb.Label(container, text="Name", bootstyle="secondary").grid(row=row, column=0, sticky=W, pady=(0, 4))
        self.name_var = tb.StringVar(value=(client.get("name") if client else ""))
        tb.Entry(container, textvariable=self.name_var, width=42).grid(
            row=row + 1,
            column=0,
            sticky=EW,
            pady=(0, 12),
        )
        row += 2

        tb.Label(container, text="Location", bootstyle="secondary").grid(
            row=row,
            column=0,
            sticky=W,
            pady=(0, 4),
        )
        self.location_var = tb.StringVar(value=(client.get("location") if client else ""))
        tb.Entry(container, textvariable=self.location_var, width=42).grid(
            row=row + 1,
            column=0,
            sticky=EW,
            pady=(0, 12),
        )
        row += 2

        self.item_var = tb.StringVar(value=(client.get("item") if client else ""))
        if self._show_item:
            tb.Label(container, text="Item", bootstyle="secondary").grid(
                row=row,
                column=0,
                sticky=W,
                pady=(0, 4),
            )
            tb.Entry(container, textvariable=self.item_var, width=42).grid(
                row=row + 1,
                column=0,
                sticky=EW,
                pady=(0, 12),
            )
            row += 2

        self.balance_var = tb.StringVar(value=(f"{client.get('balance', 0):.2f}" if client else "0.00"))
        if self._show_balance:
            tb.Label(container, text="Balance", bootstyle="secondary").grid(
                row=row,
                column=0,
                sticky=W,
                pady=(0, 4),
            )
            tb.Entry(
                container,
                textvariable=self.balance_var,
                width=42,
                validate="key",
                validatecommand=self._numeric_vcmd,
            ).grid(row=row + 1, column=0, sticky=EW, pady=(0, 16))
            row += 2

        actions = tb.Frame(container)
        actions.grid(row=row, column=0, sticky=EW)
        tb.Button(actions, text="Cancel", bootstyle="secondary-outline", command=self.destroy).pack(side=RIGHT)
        tb.Button(actions, text="Save", bootstyle="primary", command=self._submit).pack(side=RIGHT, padx=(0, 8))

        self.bind("<Return>", lambda _: self._submit())
        self.name_var.set(self.name_var.get().strip())
        self.location_var.set(self.location_var.get().strip())
        self.item_var.set(self.item_var.get().strip())
        _center_when_ready(self, parent)

    def _submit(self) -> None:
        success, message = self._on_submit(
            self.client_id,
            self.name_var.get(),
            self.location_var.get(),
            self.item_var.get(),
            self.balance_var.get(),
        )
        if not success:
            messagebox.showerror("Validation", message, parent=self)
            return
        self.destroy()


class EULADialog(tb.Toplevel):
    def __init__(self, parent) -> None:
        super().__init__(parent)
        self.title("End User License Agreement (EULA)")
        self.geometry("900x640")
        self.minsize(760, 500)
        self.transient(parent)
        self.grab_set()
        self.accepted = False

        container = tb.Frame(self, padding=16)
        container.pack(fill=BOTH, expand=True)
        container.grid_columnconfigure(0, weight=1)
        container.grid_rowconfigure(2, weight=1)

        tb.Label(
            container,
            text="END USER LICENSE AGREEMENT (EULA)",
            font=("Segoe UI", 14, "bold"),
            bootstyle="primary",
        ).grid(row=0, column=0, sticky=W)
        tb.Label(
            container,
            text="Payment Management System",
            font=("Segoe UI", 10),
            bootstyle="secondary",
        ).grid(row=1, column=0, sticky=W, pady=(0, 10))

        text_wrap = tb.Frame(container)
        text_wrap.grid(row=2, column=0, sticky="nsew")

        scrollbar = tb.Scrollbar(text_wrap, orient="vertical")
        scrollbar.pack(side=RIGHT, fill=Y)

        eula_text = tk.Text(
            text_wrap,
            wrap="word",
            yscrollcommand=scrollbar.set,
            font=("Segoe UI", 10),
            padx=12,
            pady=12,
            relief="solid",
            borderwidth=1,
        )
        eula_text.pack(side=LEFT, fill=BOTH, expand=True)
        scrollbar.configure(command=eula_text.yview)

        eula_text.insert(
            "1.0",
            """1. Agreement

This End User License Agreement ("Agreement") is a legal agreement between the Software Developer and the Client/User for the use of the Payment Management System software ("Software"). By purchasing, installing, or using the Software, the Client agrees to the terms of this Agreement.

2. Grant of License

The Developer grants the Client a perpetual, non-exclusive license to use the Software within their organization. The Client is granted the right to install and operate the Software for their internal business operations.

3. Ownership

Upon purchase, the Client owns the installed copy of the Software and all data generated within their system. However, the intellectual property rights, source code, design, and architecture of the Software remain the exclusive property of the Developer.

This Agreement does not transfer ownership of the Software's intellectual property to the Client.

4. Developer Rights

The Developer retains the right to:

- Modify and improve the Software
- Provide updates and maintenance
- Sell, license, or distribute the Software to other clients or organizations
- Continue development of future versions of the Software

5. Updates and Maintenance

The Developer will be responsible for maintaining and updating the Software when necessary. Updates may include bug fixes, improvements, and feature enhancements.

Update schedules and support terms may be agreed upon separately between the Developer and the Client.

6. Restrictions

The Client agrees not to:

- Resell, redistribute, or sublicense the Software
- Reverse engineer, modify, or alter the Software without permission from the Developer
- Copy the Software for commercial redistribution
- Claim authorship or ownership of the Software's intellectual property

7. Data Responsibility

All data stored in the system belongs to the Client. The Client is responsible for maintaining backups of their data. The Developer is not liable for data loss caused by hardware failure, misuse, or unauthorized modifications.

8. Limitation of Liability

The Software is provided "as is". The Developer shall not be held liable for any financial loss, operational disruption, or damages resulting from the use or inability to use the Software.

9. Termination

This Agreement will terminate automatically if the Client violates any of the terms stated above. Upon termination, the Client must discontinue use of the Software.

10. Acceptance

By installing, purchasing, or using the Software, the Client acknowledges that they have read, understood, and agreed to the terms of this Agreement.

Developer
Larry Kift Diolazo
""",
        )
        eula_text.configure(state="disabled")

        actions = tb.Frame(container)
        actions.grid(row=3, column=0, sticky=EW, pady=(12, 0))
        tb.Button(
            actions,
            text="Decline",
            bootstyle="danger-outline",
            command=self._decline,
        ).pack(side=RIGHT)
        tb.Button(
            actions,
            text="Accept",
            bootstyle="success",
            command=self._accept,
        ).pack(side=RIGHT, padx=(0, 8))

        self.protocol("WM_DELETE_WINDOW", self._decline)
        _center_when_ready(self, parent)

    def _accept(self) -> None:
        self.accepted = True
        self.destroy()

    def _decline(self) -> None:
        self.accepted = False
        self.destroy()


class PaymentDialog(tb.Toplevel):
    def __init__(self, parent, client_name: str, on_submit, item: str = "", existing_items: list | None = None, get_payments_for_item=None) -> None:
        super().__init__(parent)
        self.title(f"Add Payment - {client_name}")
        self.geometry("620x700")
        self.minsize(520, 560)
        self.resizable(True, True)
        self.transient(parent)
        self.grab_set()

        self._on_submit = on_submit
        self._numeric_vcmd = (self.register(_is_valid_numeric_input), "%P")
        self._payment_entries: list[tuple[str, tb.StringVar, tb.Entry]] = []
        self._month_names = list(calendar.month_name)[1:]
        self._existing_items: list[str] = existing_items or []
        self._get_payments_for_item = get_payments_for_item  # callable(item) -> list[dict]

        container = tb.Frame(self, padding=18)
        container.pack(fill=BOTH, expand=True)

        header = tb.Frame(container)
        header.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        header.grid_columnconfigure(3, weight=1)

        tb.Label(header, text="Payment Calendar", font=("Segoe UI", 12, "bold")).grid(row=0, column=0, sticky=W)

        now = datetime.now()
        self.month_var = tb.StringVar(value=self._month_names[now.month - 1])
        self.year_var = tb.StringVar(value=str(now.year))

        month_combo = tb.Combobox(
            header,
            textvariable=self.month_var,
            values=self._month_names,
            state="readonly",
            width=12,
        )
        month_combo.grid(row=0, column=1, padx=(12, 0))
        year_combo = tb.Combobox(
            header,
            textvariable=self.year_var,
            values=[str(now.year - 1), str(now.year), str(now.year + 1), str(now.year + 2)],
            state="readonly",
            width=6,
        )
        year_combo.grid(row=0, column=2, padx=(8, 0))
        month_combo.bind("<<ComboboxSelected>>", lambda _e: self._rebuild_calendar())
        year_combo.bind("<<ComboboxSelected>>", lambda _e: self._rebuild_calendar())

        table_wrap = tb.Frame(container)
        table_wrap.grid(row=1, column=0, sticky="nsew", pady=(0, 12))
        container.grid_rowconfigure(1, weight=1)
        container.grid_columnconfigure(0, weight=1)

        canvas = tk.Canvas(table_wrap, highlightthickness=0)
        vscroll = tb.Scrollbar(table_wrap, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vscroll.set)

        self.calendar_inner = tb.Frame(canvas)
        self.calendar_inner.bind("<Configure>", lambda _e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self.calendar_inner, anchor="nw")

        canvas.grid(row=0, column=0, sticky="nsew")
        vscroll.grid(row=0, column=1, sticky="ns")
        table_wrap.grid_rowconfigure(0, weight=1)
        table_wrap.grid_columnconfigure(0, weight=1)
        self._bind_mousewheel(canvas)

        style = tb.Style()
        style.configure("PaymentEmpty.TEntry", fieldbackground="#fecaca")
        style.configure("PaymentFilled.TEntry", fieldbackground="#dcfce7")

        # Item field — Combobox if items available, else plain Entry
        tb.Label(container, text="Item", bootstyle="secondary").grid(row=2, column=0, sticky=W, pady=(0, 4))
        self.item_var = tb.StringVar(value=item)
        if self._existing_items:
            item_combo = tb.Combobox(
                container,
                textvariable=self.item_var,
                values=self._existing_items,
                state="readonly",
                width=42,
            )
            item_combo.grid(row=3, column=0, sticky=EW, pady=(0, 12))
            # When item changes, reload the calendar with that item's payments
            self.item_var.trace_add("write", lambda *_: self._rebuild_calendar())
        else:
            tb.Entry(container, textvariable=self.item_var, width=42).grid(row=3, column=0, sticky=EW, pady=(0, 12))

        tb.Label(container, text="Note", bootstyle="secondary").grid(row=4, column=0, sticky=W, pady=(0, 4))
        self.note_var = tb.StringVar(value="")
        tb.Entry(container, textvariable=self.note_var, width=42).grid(row=5, column=0, sticky=EW, pady=(0, 16))

        actions = tb.Frame(container)
        actions.grid(row=6, column=0, sticky=EW)
        tb.Button(actions, text="Cancel", bootstyle="secondary-outline", command=self.destroy).pack(side=RIGHT)
        tb.Button(actions, text="Record", bootstyle="success", command=self._submit).pack(side=RIGHT, padx=(0, 8))

        self.bind("<Return>", lambda _: self._submit())
        self._rebuild_calendar()
        _center_when_ready(self, parent)

    def _submit(self) -> None:
        payments: list[tuple[str, str]] = []
        for date_value, var, _entry in self._payment_entries:
            value = str(var.get()).strip()
            if not value:
                continue
            try:
                parsed = float(value.replace(",", ""))
            except ValueError:
                messagebox.showerror("Validation", f"Invalid amount for {date_value}.", parent=self)
                return
            if parsed <= 0:
                messagebox.showerror("Validation", f"Amount must be greater than 0 for {date_value}.", parent=self)
                return
            payments.append((date_value, value))

        if not payments:
            messagebox.showerror("Validation", "Enter at least one payment amount.", parent=self)
            return

        success, message = self._on_submit(
            payments,
            self.item_var.get(),
            self.note_var.get(),
        )
        if not success:
            messagebox.showerror("Validation", message, parent=self)
            return
        messagebox.showinfo("Saved", message, parent=self)
        self.destroy()

    def _rebuild_calendar(self) -> None:
        for child in self.calendar_inner.winfo_children():
            child.destroy()
        self._payment_entries.clear()

        tb.Label(self.calendar_inner, text="Date", bootstyle="secondary").grid(
            row=0, column=0, sticky=W, pady=(0, 6)
        )
        tb.Label(self.calendar_inner, text="Payment", bootstyle="secondary").grid(
            row=0, column=1, sticky=W, pady=(0, 6)
        )

        year = int(self.year_var.get())
        month_name = self.month_var.get()
        month = self._month_names.index(month_name) + 1
        days_in_month = calendar.monthrange(year, month)[1]

        # Build a date->amount map from existing payments for this item+month
        existing_by_date: dict[str, float] = {}
        if self._get_payments_for_item is not None:
            selected_item = self.item_var.get().strip()
            try:
                payments = self._get_payments_for_item(selected_item)
                for p in payments:
                    raw_date = str(p.get("created_at") or "").strip()[:10]  # "YYYY-MM-DD"
                    try:
                        pd = datetime.strptime(raw_date, "%Y-%m-%d")
                    except ValueError:
                        continue
                    if pd.year == year and pd.month == month:
                        existing_by_date[raw_date] = (
                            existing_by_date.get(raw_date, 0.0) + float(p.get("amount") or 0.0)
                        )
            except Exception:
                pass

        for idx in range(1, days_in_month + 1):
            date_value = datetime(year, month, idx).strftime("%Y-%m-%d")
            tb.Label(self.calendar_inner, text=date_value).grid(
                row=idx, column=0, sticky=W, pady=2, padx=(0, 10)
            )

            existing_amount = existing_by_date.get(date_value, 0.0)
            initial_value = f"{existing_amount:.2f}" if existing_amount > 0 else ""

            amount_var = tb.StringVar(value=initial_value)
            entry = tb.Entry(
                self.calendar_inner,
                textvariable=amount_var,
                width=20,
                validate="key",
                validatecommand=self._numeric_vcmd,
                style="PaymentFilled.TEntry" if initial_value else "PaymentEmpty.TEntry",
            )
            entry.grid(row=idx, column=1, sticky=W, pady=2)
            amount_var.trace_add("write", lambda *_args, var=amount_var, ent=entry: self._update_entry_style(var, ent))
            self._payment_entries.append((date_value, amount_var, entry))

    @staticmethod
    def _bind_mousewheel(canvas: tk.Canvas) -> None:
        def on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        def _bind(_event):
            canvas.bind_all("<MouseWheel>", on_mousewheel)

        def _unbind(_event):
            canvas.unbind_all("<MouseWheel>")

        canvas.bind("<Enter>", _bind)
        canvas.bind("<Leave>", _unbind)

    @staticmethod
    def _update_entry_style(var: tb.StringVar, entry: tb.Entry) -> None:
        value = str(var.get()).strip()
        if not value:
            entry.configure(style="PaymentEmpty.TEntry")
            return
        try:
            parsed = float(value.replace(",", ""))
        except ValueError:
            entry.configure(style="PaymentEmpty.TEntry")
            return
        entry.configure(style="PaymentFilled.TEntry" if parsed > 0 else "PaymentEmpty.TEntry")


class AddBalanceDialog(tb.Toplevel):
    def __init__(
        self,
        parent,
        client_name: str,
        on_submit,
        item: str = "",
        existing_items: list[str] | None = None,
        excess_amount: float = 0.0,
    ) -> None:
        super().__init__(parent)
        self.title(f"Add Balance - {client_name}")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()

        self._on_submit = on_submit
        self._numeric_vcmd = (self.register(_is_valid_numeric_input), "%P")

        container = tb.Frame(self, padding=18)
        container.pack(fill=BOTH, expand=True)

        tb.Label(container, text="Additional Balance", bootstyle="secondary").pack(anchor=W, pady=(0, 4))
        self.amount_var = tb.StringVar(value="")
        tb.Entry(
            container,
            textvariable=self.amount_var,
            width=42,
            validate="key",
            validatecommand=self._numeric_vcmd,
        ).pack(fill=X, pady=(0, 16))

        tb.Label(container, text="Item", bootstyle="secondary").pack(anchor=W, pady=(0, 4))
        self.item_var = tb.StringVar(value=item)
        tb.Entry(container, textvariable=self.item_var, width=42).pack(fill=X, pady=(0, 16))

        self.excess_target_var = tb.StringVar(value="")
        normalized_existing = [str(value).strip() for value in (existing_items or []) if str(value).strip()]
        if excess_amount > 0 and normalized_existing:
            tb.Label(
                container,
                text=f"Excess available: {excess_amount:.2f}",
                bootstyle="warning",
            ).pack(anchor=W, pady=(0, 4))
            tb.Label(container, text="Apply Excess To", bootstyle="secondary").pack(anchor=W, pady=(0, 4))
            self.excess_target_combo = tb.Combobox(
                container,
                textvariable=self.excess_target_var,
                values=normalized_existing,
                state="readonly",
                width=40,
            )
            if item and item in normalized_existing:
                self.excess_target_var.set(item)
            else:
                self.excess_target_var.set(normalized_existing[0])
            self.excess_target_combo.pack(fill=X, pady=(0, 16))
        else:
            self.excess_target_combo = None

        actions = tb.Frame(container)
        actions.pack(fill=X)
        tb.Button(actions, text="Cancel", bootstyle="secondary-outline", command=self.destroy).pack(side=RIGHT)
        tb.Button(actions, text="Apply", bootstyle="warning", command=self._submit).pack(side=RIGHT, padx=(0, 8))

        self.bind("<Return>", lambda _: self._submit())
        _center_when_ready(self, parent)

    def _submit(self) -> None:
        if not self._is_positive_amount(self.amount_var.get()):
            messagebox.showerror(
                "Validation",
                "Additional balance must be greater than 0. Negative numbers are not allowed.",
                parent=self,
            )
            return
        success, message = self._on_submit(
            self.amount_var.get(),
            self.item_var.get(),
            self.excess_target_var.get().strip(),
        )
        if not success:
            messagebox.showerror("Validation", message, parent=self)
            return
        messagebox.showinfo("Updated", message, parent=self)
        self.destroy()

    @staticmethod
    def _is_positive_amount(value: str) -> bool:
        try:
            return float(str(value).replace(",", "").strip()) > 0
        except (TypeError, ValueError):
            return False


class EditItemDialog(tb.Toplevel):
    def __init__(self, parent, current_item: str, on_submit) -> None:
        super().__init__(parent)
        self.title("Edit Item")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()

        self._on_submit = on_submit

        container = tb.Frame(self, padding=18)
        container.pack(fill=BOTH, expand=True)

        current_label = current_item if current_item else "Unspecified"
        tb.Label(container, text="Current Item", bootstyle="secondary").grid(row=0, column=0, sticky=W, pady=(0, 4))
        tb.Label(container, text=current_label, font=("Segoe UI", 11, "bold")).grid(
            row=1, column=0, sticky=W, pady=(0, 12)
        )

        tb.Label(container, text="New Item Name", bootstyle="secondary").grid(row=2, column=0, sticky=W, pady=(0, 4))
        self.item_var = tb.StringVar(value=current_item)
        tb.Entry(container, textvariable=self.item_var, width=42).grid(row=3, column=0, sticky=EW, pady=(0, 16))

        tb.Label(container, text="Balance", bootstyle="secondary").grid(row=4, column=0, sticky=W, pady=(0, 4))
        self.balance_var = tb.StringVar(value="")
        tb.Entry(
            container,
            textvariable=self.balance_var,
            width=42,
            validate="key",
            validatecommand=(self.register(_is_valid_numeric_input), "%P"),
        ).grid(row=5, column=0, sticky=EW, pady=(0, 16))

        actions = tb.Frame(container)
        actions.grid(row=6, column=0, sticky=EW)
        tb.Button(actions, text="Cancel", bootstyle="secondary-outline", command=self.destroy).pack(side=RIGHT)
        tb.Button(actions, text="Save", bootstyle="primary", command=self._submit).pack(side=RIGHT, padx=(0, 8))

        self.bind("<Return>", lambda _: self._submit())
        _center_when_ready(self, parent)

    def _submit(self) -> None:
        success, message = self._on_submit(self.item_var.get(), self.balance_var.get())
        if not success:
            messagebox.showerror("Validation", message, parent=self)
            return
        if message:
            messagebox.showinfo("Updated", message, parent=self)
        self.destroy()


class ReturnItemDialog(tb.Toplevel):
    def __init__(self, parent, client_name: str, items: list[str]) -> None:
        super().__init__(parent)
        self.title(f"Return Item - {client_name}")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()
        self.result: tuple[str, str] | None = None  # (item, note)

        container = tb.Frame(self, padding=18)
        container.pack(fill=BOTH, expand=True)

        tb.Label(
            container,
            text=f"Return item for: {client_name}",
            font=("Segoe UI", 11, "bold"),
            bootstyle="danger",
        ).grid(row=0, column=0, columnspan=2, sticky=W, pady=(0, 6))

        tb.Label(
            container,
            text="The full unpaid balance of the item will be deducted automatically.",
            bootstyle="secondary",
        ).grid(row=1, column=0, columnspan=2, sticky=W, pady=(0, 14))

        tb.Label(container, text="Item", bootstyle="secondary").grid(row=2, column=0, sticky=W, pady=(0, 4))
        self.item_var = tb.StringVar()
        item_combo = tb.Combobox(
            container,
            textvariable=self.item_var,
            values=items,
            state="readonly",
            width=38,
        )
        item_combo.grid(row=3, column=0, columnspan=2, sticky=EW, pady=(0, 12))
        if items:
            item_combo.current(0)

        tb.Label(container, text="Note (optional)", bootstyle="secondary").grid(row=4, column=0, sticky=W, pady=(0, 4))
        self.note_var = tb.StringVar()
        tb.Entry(container, textvariable=self.note_var, width=42).grid(
            row=5, column=0, columnspan=2, sticky=EW, pady=(0, 18)
        )

        actions = tb.Frame(container)
        actions.grid(row=6, column=0, columnspan=2, sticky=EW)
        tb.Button(actions, text="Cancel", bootstyle="secondary-outline", command=self.destroy).pack(side=RIGHT)
        tb.Button(
            actions,
            text="Confirm Return",
            bootstyle="danger",
            command=self._submit,
        ).pack(side=RIGHT, padx=(0, 8))

        container.columnconfigure(0, weight=1)
        self.bind("<Return>", lambda _: self._submit())
        _center_when_ready(self, parent)

    def _submit(self) -> None:
        item = self.item_var.get().strip()
        if not item:
            messagebox.showerror("Validation", "Please select an item.", parent=self)
            return
        self.result = (item, self.note_var.get().strip())
        self.destroy()


class ClientInfoDialog(tb.Toplevel):
    def __init__(
        self,
        parent,
        client: dict,
        client_controller,
        payment_controller,
        on_refresh,
        on_close=None,
    ) -> None:
        super().__init__(parent)
        self.client_id = client.get("id")
        self.client_controller = client_controller
        self.payment_controller = payment_controller
        self.on_refresh = on_refresh
        self.on_close = on_close
        self._closed = False
        self._placeholder_item = "No items yet"

        self.title(f"Client Info - {client.get('name', 'Client')}")
        self.geometry("1060x700")
        self.minsize(900, 600)
        self.transient(parent)
        self.grab_set()

        container = tb.Frame(self, padding=16)
        container.pack(fill=BOTH, expand=True)
        container.grid_columnconfigure(0, weight=1)
        container.grid_rowconfigure(2, weight=1)

        header = tb.Frame(container, padding=12, bootstyle="light")
        header.grid(row=0, column=0, sticky="ew")
        header.grid_columnconfigure(0, weight=1)

        self.name_var = tb.StringVar(value="")
        self.meta_var = tb.StringVar(value="")
        self.balance_var = tb.StringVar(value="")
        self.excess_var = tb.StringVar(value="")
        self.barcode_var = tb.StringVar(value="")
        self._barcode_image = None  # PIL ImageTk reference

        tb.Label(header, textvariable=self.name_var, font=("Segoe UI", 16, "bold"), bootstyle="primary").grid(
            row=0, column=0, sticky="w"
        )
        tb.Label(header, textvariable=self.meta_var, bootstyle="secondary").grid(
            row=1, column=0, sticky="w", pady=(4, 0)
        )

        stats = tb.Frame(header, bootstyle="light")
        stats.grid(row=0, column=1, rowspan=2, sticky="e")
        tb.Label(stats, text="Current Balance", bootstyle="secondary").grid(row=0, column=0, sticky="e")
        tb.Label(stats, textvariable=self.balance_var, font=("Segoe UI", 12, "bold")).grid(
            row=1, column=0, sticky="e"
        )
        tb.Label(stats, text="Excess Payment", bootstyle="secondary").grid(row=0, column=1, sticky="e", padx=(18, 0))
        tb.Label(stats, textvariable=self.excess_var, font=("Segoe UI", 12, "bold")).grid(
            row=1, column=1, sticky="e", padx=(18, 0)
        )

        # ── Body: barcode panel (left) + summary table (right) ────────────────
        body = tb.Frame(container)
        body.grid(row=2, column=0, sticky="nsew", pady=(14, 0))
        body.grid_columnconfigure(1, weight=1)
        body.grid_rowconfigure(0, weight=1)        # Barcode side panel — fixed width, scrollable if needed
        barcode_card = tb.Frame(body, padding=12, bootstyle="light", width=260)
        barcode_card.grid(row=0, column=0, sticky="ns", padx=(0, 12))
        barcode_card.grid_propagate(False)

        tb.Label(barcode_card, text="Barcode", font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0, 6))
        self.barcode_image_label = tb.Label(barcode_card, text="No barcode image", bootstyle="secondary")
        self.barcode_image_label.pack(anchor="center", pady=(0, 6))
        self.barcode_code_label = tb.Label(
            barcode_card, textvariable=self.barcode_var, bootstyle="secondary",
            wraplength=236, justify="center",
        )
        self.barcode_code_label.pack(anchor="center")

        summary_card = tb.Frame(body, padding=12, bootstyle="light")
        summary_card.grid(row=0, column=1, sticky="nsew")
        summary_card.grid_columnconfigure(0, weight=1)
        summary_card.grid_rowconfigure(1, weight=1)

        tb.Label(summary_card, text="Balances per Item", font=("Segoe UI", 12, "bold")).grid(
            row=0, column=0, sticky="w", pady=(0, 6)
        )

        table_wrap = tb.Frame(summary_card)
        table_wrap.grid(row=1, column=0, sticky="nsew")
        table_wrap.grid_columnconfigure(0, weight=1)
        table_wrap.grid_rowconfigure(0, weight=1)

        style = tb.Style()
        style.configure("ClientInfo.Treeview", rowheight=26, font=("Segoe UI", 10))
        style.configure("ClientInfo.Treeview.Heading", font=("Segoe UI", 10, "bold"))

        self.summary_table = tb.Treeview(
            table_wrap,
            columns=("item", "balance", "paid", "remaining"),
            show="headings",
            style="ClientInfo.Treeview",
        )
        self.summary_table.heading("item", text="Item", anchor="w")
        self.summary_table.heading("balance", text="Total Balance", anchor="center")
        self.summary_table.heading("paid", text="Paid Balance", anchor="center")
        self.summary_table.heading("remaining", text="Remaining", anchor="center")
        self.summary_table.column("item", width=360, minwidth=220, anchor="w", stretch=True)
        self.summary_table.column("balance", width=160, minwidth=120, anchor="center", stretch=True)
        self.summary_table.column("paid", width=160, minwidth=120, anchor="center", stretch=True)
        self.summary_table.column("remaining", width=160, minwidth=120, anchor="center", stretch=True)
        self.summary_table.tag_configure("paid", background="#dcfce7", foreground="#14532d")
        self.summary_table.tag_configure("returned", background="#fee2e2", foreground="#7f1d1d")

        yscroll = tb.Scrollbar(table_wrap, orient="vertical", command=self.summary_table.yview)
        self.summary_table.configure(yscrollcommand=yscroll.set)
        self.summary_table.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")

        actions = tb.Frame(container)
        actions.grid(row=3, column=0, sticky="ew", pady=(14, 0))

        # Row 1: primary actions
        row1 = tb.Frame(actions)
        row1.pack(fill=X, anchor="w")
        tb.Button(row1, text="Add Payment", bootstyle="success", command=self._add_payment).pack(side=LEFT)
        tb.Button(row1, text="Add Balance", bootstyle="warning", command=self._add_balance).pack(side=LEFT, padx=(8, 0))
        tb.Button(row1, text="Export Barcode", bootstyle="info-outline", command=self._export_barcode_image).pack(side=LEFT, padx=(8, 0))
        tb.Button(row1, text="Payment History", bootstyle="secondary-outline", command=self._open_payment_history).pack(side=LEFT, padx=(8, 0))

        # Row 2: secondary actions
        row2 = tb.Frame(actions)
        row2.pack(fill=X, anchor="w", pady=(6, 0))
        tb.Button(row2, text="Return Item", bootstyle="danger-outline", command=self._return_item).pack(side=LEFT)
        tb.Button(row2, text="Edit Item", bootstyle="secondary", command=self._edit_item).pack(side=LEFT, padx=(8, 0))
        tb.Button(row2, text="Edit Client", bootstyle="primary", command=self._edit_client).pack(side=LEFT, padx=(8, 0))
        tb.Button(row2, text="Delete Client", bootstyle="danger", command=self._delete_client).pack(side=LEFT, padx=(8, 0))
        tb.Button(row2, text="Close", bootstyle="secondary-outline", command=self._handle_close).pack(side=RIGHT)

        self._load_client()
        self._load_summary()
        self.protocol("WM_DELETE_WINDOW", self._handle_close)
        self.bind("<Destroy>", self._on_destroy, add="+")
        _center_when_ready(self, parent)

    def refresh(self) -> None:
        self._load_client()
        self._load_summary()

    def _handle_close(self) -> None:
        if self._closed:
            return
        self._closed = True
        if callable(self.on_close):
            self.on_close(self.client_id)
        try:
            self.grab_release()
        except tk.TclError:
            pass
        self.destroy()

    def _on_destroy(self, _event=None) -> None:
        if self._closed:
            return
        self._closed = True
        if callable(self.on_close):
            self.on_close(self.client_id)

    def _load_client(self) -> None:
        client = self.client_controller.get_client(self.client_id)
        if not client:
            messagebox.showerror("Not found", "Client no longer exists.", parent=self)
            self._handle_close()
            return

        self.client = client
        location = str(client.get("location") or "").strip() or "-"
        item = str(client.get("item") or "").strip() or "-"
        self.name_var.set(client.get("name") or "Client")
        self.meta_var.set(f"Location: {location} | Default item: {item}")
        self.balance_var.set(f"{float(client.get('balance', 0.0)):.2f}")
        self.excess_var.set(f"{float(client.get('excess_payment', 0.0)):.2f}")
        barcode_value = str(client.get("barcode_value") or "").strip()
        self.barcode_var.set(f"Barcode: {barcode_value}" if barcode_value else "Barcode: -")
        self._load_barcode_image(client.get("barcode_image_path"))

    def _load_barcode_image(self, barcode_path: str | None) -> None:
        path_text = str(barcode_path or "").strip()
        if not path_text:
            self._barcode_image = None
            self.barcode_image_label.configure(image="", text="No barcode image")
            return

        image_path = Path(path_text)
        if not image_path.exists():
            self._barcode_image = None
            self.barcode_image_label.configure(image="", text="No barcode image")
            return

        try:
            from PIL import Image, ImageTk
            img = Image.open(str(image_path)).convert("RGB")
            # Fit inside the 236×160 panel while preserving aspect ratio
            max_w, max_h = 236, 160
            img.thumbnail((max_w, max_h), Image.LANCZOS)
            self._barcode_image = ImageTk.PhotoImage(img)
            self.barcode_image_label.configure(image=self._barcode_image, text="")
        except Exception:
            self._barcode_image = None
            self.barcode_image_label.configure(image="", text="No barcode image")

    def _item_key(self, value: str | None) -> str:
        normalized = str(value or "").strip()
        return normalized if normalized else "Unspecified"

    def _load_summary(self) -> None:
        for item in self.summary_table.get_children():
            self.summary_table.delete(item)

        events = self.payment_controller.list_client_balance_events(self.client_id)
        payments = self.payment_controller.list_client_payments(self.client_id)
        returned_items = self.payment_controller.list_returned_items(self.client_id)

        summary: dict[str, dict[str, float]] = {}
        for event in events:
            key = self._item_key(event.get("item"))
            entry = summary.setdefault(key, {"balance": 0.0, "paid": 0.0, "returned": 0.0})
            entry["balance"] += float(event.get("amount", 0.0) or 0.0)
            entry["paid"] += float(event.get("applied_excess", 0.0) or 0.0)

        for payment in payments:
            key = self._item_key(payment.get("item"))
            entry = summary.setdefault(key, {"balance": 0.0, "paid": 0.0, "returned": 0.0})
            entry["paid"] += float(payment.get("amount", 0.0) or 0.0)

        for ret in returned_items:
            key = self._item_key(ret.get("item"))
            entry = summary.setdefault(key, {"balance": 0.0, "paid": 0.0, "returned": 0.0})
            entry["returned"] += float(ret.get("amount", 0.0) or 0.0)

        remaining_total = 0.0
        visible_rows = 0

        for key in sorted(summary.keys(), key=lambda value: value.lower()):
            balance = summary[key]["balance"]
            paid = summary[key]["paid"]
            returned = summary[key].get("returned", 0.0)
            remaining = max(0.0, balance - paid - returned)
            if remaining <= 1e-6:
                continue

            remaining_total += remaining
            if returned > 0:
                tags = ("returned",)
                remaining_display = f"{remaining:.2f} (partial return)"
            else:
                tags = ()
                remaining_display = f"{remaining:.2f}"

            self.summary_table.insert(
                "",
                "end",
                values=(
                    key,
                    f"{balance:.2f}",
                    f"{paid:.2f}",
                    remaining_display,
                ),
                tags=tags,
            )
            visible_rows += 1

        if visible_rows == 0:
            self.summary_table.insert(
                "",
                "end",
                values=(self._placeholder_item, "-", "-", "-"),
            )

        self.balance_var.set(f"{remaining_total:.2f}")

    def _selected_item(self) -> str:
        selection = self.summary_table.selection()
        if not selection:
            return ""
        item_label = str(self.summary_table.item(selection[0], "values")[0] or "").strip()
        if not item_label or item_label == self._placeholder_item:
            return ""
        return "" if item_label == "Unspecified" else item_label

    def _selected_item_label(self) -> str:
        selection = self.summary_table.selection()
        if not selection:
            return ""
        item_label = str(self.summary_table.item(selection[0], "values")[0] or "").strip()
        if not item_label or item_label == self._placeholder_item:
            return ""
        return item_label

    def _add_payment(self) -> None:
        item = self._selected_item()
        rows = self.summary_table.get_children()
        valid_rows = [
            row
            for row in rows
            if str(self.summary_table.item(row, "values")[0]) != self._placeholder_item
        ]
        if not item:
            if len(valid_rows) == 1:
                first = self.summary_table.item(valid_rows[0], "values")[0]
                item = "" if first == "Unspecified" else str(first)
            elif len(valid_rows) > 1:
                messagebox.showwarning(
                    "Select item",
                    "Select an item row before recording a payment.",
                    parent=self,
                )
                return

        # Collect all item names for the dropdown
        existing_items: list[str] = []
        for row_id in valid_rows:
            label = str(self.summary_table.item(row_id, "values")[0] or "").strip()
            if label and label != self._placeholder_item and label != "Unspecified":
                existing_items.append(label)
        existing_items = sorted(set(existing_items), key=lambda v: v.lower())

        def get_payments_for_item(item_value: str):
            return self.payment_controller.list_client_payments(self.client_id)

        def on_submit(payments, item_value, note):
            success, message = self.payment_controller.create_calendar_payments(
                self.client_id,
                item_value,
                note,
                payments,
            )
            if success:
                self.on_refresh()
                self._load_client()
                self._load_summary()
            return success, message

        PaymentDialog(
            self,
            self.client.get("name", "Client"),
            on_submit,
            item=item,
            existing_items=existing_items,
            get_payments_for_item=get_payments_for_item,
        )

    def _add_balance(self) -> None:
        item = self._selected_item()
        existing_item_labels: list[str] = []
        for row_id in self.summary_table.get_children():
            values = self.summary_table.item(row_id, "values")
            label = str(values[0] or "").strip()
            if not label or label == self._placeholder_item:
                continue
            remaining_text = str(values[3] or "").strip()
            numeric_token = remaining_text.split(" ")[0] if remaining_text else "0"
            try:
                remaining_value = float(numeric_token.replace(",", ""))
            except ValueError:
                remaining_value = 0.0
            if remaining_value <= 1e-6:
                continue
            existing_item_labels.append(label)

        existing_items = sorted(set(existing_item_labels), key=lambda value: value.lower())

        def on_submit(amount, item_value, excess_target_item):
            normalized_target = str(excess_target_item or "").strip()
            if normalized_target == "Unspecified":
                normalized_target = ""
            success, message = self.client_controller.increase_balance(
                self.client_id,
                amount,
                item_value,
                excess_target_item=normalized_target or None,
            )
            if success:
                self.on_refresh()
                self._load_client()
                self._load_summary()
            return success, message

        AddBalanceDialog(
            self,
            self.client.get("name", "Client"),
            on_submit,
            item=item,
            existing_items=existing_items,
            excess_amount=float(self.client.get("excess_payment", 0.0) or 0.0),
        )

    def _open_payment_history(self) -> None:
        grouped_history = self.payment_controller.list_client_payments_by_balance(self.client_id)
        PaymentHistoryDialog(self, self.client.get("name", "Client"), grouped_history)

    def _export_barcode_image(self) -> None:
        current_client = self.client_controller.get_client(self.client_id)
        if not current_client:
            messagebox.showerror("Not found", "Client no longer exists.", parent=self)
            return

        barcode_path_text = str(current_client.get("barcode_image_path") or "").strip()
        if not barcode_path_text:
            messagebox.showerror(
                "Barcode Unavailable",
                "Barcode image is not available. Please install barcode dependencies and reopen the app.",
                parent=self,
            )
            return

        barcode_path = Path(barcode_path_text)
        if not barcode_path.exists():
            messagebox.showerror(
                "Barcode Missing",
                f"Barcode image file not found:\n{barcode_path}",
                parent=self,
            )
            return

        safe_name = "".join(ch for ch in str(current_client.get("name") or "client") if ch not in '\\/:*?"<>|').strip()
        default_name = f"{safe_name or 'client'}_barcode.png"
        save_path = filedialog.asksaveasfilename(
            parent=self,
            title="Export Barcode Image",
            defaultextension=".png",
            initialfile=default_name,
            filetypes=[("PNG Image", "*.png"), ("All files", "*.*")],
        )
        if not save_path:
            return

        try:
            shutil.copy2(barcode_path, save_path)
        except Exception as exc:
            messagebox.showerror("Export Error", f"Could not export barcode image:\n{exc}", parent=self)
            return
        messagebox.showinfo("Exported", f"Barcode image exported to:\n{save_path}", parent=self)

    def _return_item(self) -> None:
        items = self.payment_controller.list_client_items(self.client_id)

        preselected = self._selected_item_label()
        if preselected and preselected != "Unspecified" and preselected not in items:
            items.insert(0, preselected)

        if not items:
            messagebox.showinfo("No Items", "This client has no items to return.", parent=self)
            return

        dialog = ReturnItemDialog(self, self.client.get("name", "Client"), items)
        if preselected and preselected in items:
            dialog.item_var.set(preselected)

        dialog.result = None
        self.wait_window(dialog)

        if not dialog.result:
            return

        item, note = dialog.result
        ok, message = self.payment_controller.return_item(
            self.client_id,
            item,
            note or "Item returned",
        )
        if ok:
            self.on_refresh()
            self._load_client()
            self._load_summary()
            messagebox.showinfo("Return Recorded", message, parent=self)
        else:
            messagebox.showerror("Error", message, parent=self)

    def _edit_item(self) -> None:
            item_label = self._selected_item_label()
            if not item_label:
                messagebox.showwarning(
                    "Select item",
                    "Select an item row to edit.",
                    parent=self,
                )
                return

            old_item = "" if item_label == "Unspecified" else item_label

            def on_submit(new_item: str, balance_text: str):
                success, message = self.payment_controller.rename_item_for_client(
                    self.client_id,
                    old_item,
                    new_item,
                )
                if not success:
                    return success, message

                balance_text = str(balance_text or "").strip()
                if balance_text:
                    try:
                        parsed_balance = float(balance_text.replace(",", ""))
                    except ValueError:
                        return False, "Balance must be a valid number."
                    if parsed_balance > 0:
                        success, balance_message = self.client_controller.increase_balance(
                            self.client_id,
                            balance_text,
                            new_item,
                        )
                        if not success:
                            return success, balance_message

                self.on_refresh()
                self._load_client()
                self._load_summary()
                return True, message

            EditItemDialog(self, old_item, on_submit)

    def _edit_client(self) -> None:
        def on_submit(client_id, name, location, item, balance):
            success, message = self.client_controller.edit_client(
                client_id,
                name,
                location,
                item,
                balance,
            )
            if success:
                self.on_refresh()
                self._load_client()
                self._load_summary()
            return success, message

        ClientFormDialog(
            self,
            "Edit Client",
            on_submit,
            self.client,
            show_item=False,
            show_balance=False,
        )

    def _delete_client(self) -> None:
        confirmed = messagebox.askyesno(
            "Delete client",
            f"Delete {self.client.get('name', 'this client')} and all their payments?",
            icon="warning",
            parent=self,
        )
        if not confirmed:
            return

        success, message = self.client_controller.remove_client(self.client_id)
        if success:
            self.on_refresh()
            messagebox.showinfo("Deleted", message, parent=self)
            self._handle_close()
            return
        messagebox.showerror("Error", message, parent=self)


class PaymentHistoryDialog(tb.Toplevel):
    def __init__(self, parent, client_name: str, grouped_history: list[dict]) -> None:
        super().__init__(parent)
        self.title(f"Payment History - {client_name}")
        self.geometry("560x760")
        self.minsize(500, 520)
        self.transient(parent)
        self.parent = parent
        self._closed = False
        self.grab_set()
        self.client_name = client_name
        self.grouped_history = grouped_history
        self.month_entries: list[dict] = self._build_month_data()
        self.month_var = tb.StringVar(value="")
        self._column_to_item: dict[str, str] = {}

        container = tb.Frame(self, padding=16)
        container.pack(fill=BOTH, expand=True)
        container.grid_columnconfigure(0, weight=1)
        container.grid_rowconfigure(3, weight=1)

        tb.Label(
            container,
            text=self.client_name,
            font=("Segoe UI", 14, "bold"),
            bootstyle="primary",
            anchor="center",
        ).grid(row=0, column=0, sticky="ew")

        picker_row = tb.Frame(container)
        picker_row.grid(row=1, column=0, sticky="ew", pady=(12, 10))
        picker_row.grid_columnconfigure(1, weight=1)
        tb.Label(picker_row, text="Month", bootstyle="secondary").grid(row=0, column=0, sticky="w")

        options = [entry["label"] for entry in self.month_entries]
        if options:
            self.month_var.set(options[-1])
        else:
            self.month_var.set(datetime.now().strftime("%B %Y"))

        self.month_combo = tb.Combobox(
            picker_row,
            textvariable=self.month_var,
            values=options,
            state="readonly",
            width=44,
        )
        self.month_combo.grid(row=0, column=1, sticky="ew", padx=(10, 0))
        self.month_combo.bind("<<ComboboxSelected>>", lambda _e: self._render_selected_month())

        table_wrap = tb.Frame(container)
        table_wrap.grid(row=3, column=0, sticky="nsew")
        table_wrap.grid_rowconfigure(0, weight=1)
        table_wrap.grid_columnconfigure(0, weight=1)

        style = tb.Style()
        style.configure("HistoryTemplate.Treeview", rowheight=26, font=("Segoe UI", 10))
        style.configure("HistoryTemplate.Treeview.Heading", font=("Segoe UI", 10, "bold"))

        self.table = tb.Treeview(
            table_wrap,
            columns=("date",),
            show="headings",
            style="HistoryTemplate.Treeview",
        )
        self.table.heading("date", text="Date", anchor="center")
        self.table.column("date", width=280, minwidth=180, anchor="center", stretch=True)

        self.table.tag_configure("odd", background="#eaf2ff")
        self.table.tag_configure("even", background="#ffffff")
        self.table.tag_configure("summary", background="#f5f5f5", font=("Segoe UI", 10, "bold"))

        yscroll = tb.Scrollbar(table_wrap, orient="vertical", command=self.table.yview)
        xscroll = tb.Scrollbar(table_wrap, orient="horizontal", command=self.table.xview)
        self.table.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)

        self.table.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")

        actions = tb.Frame(container)
        actions.grid(row=4, column=0, sticky="ew", pady=(12, 0))
        tb.Button(
            actions,
            text="Export Excel",
            bootstyle="success-outline",
            command=self._export_excel,
        ).pack(side=LEFT)
        tb.Button(actions, text="Close", bootstyle="secondary-outline", command=self._handle_close).pack(side=RIGHT)
        self._render_selected_month()
        self.protocol("WM_DELETE_WINDOW", self._handle_close)
        _center_when_ready(self, parent)

    def _export_excel(self) -> None:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror(
                "Export Error",
                "Excel export requires openpyxl. Install dependencies first.",
                parent=self,
            )
            return

        selected_month = self._selected_month_entry()
        if selected_month is None:
            messagebox.showwarning("No History", "No payment history is available for export.", parent=self)
            return

        safe_name = "".join(ch for ch in self.client_name if ch not in '\\/:*?"<>|').strip() or "client"
        month_slug = selected_month["key"]
        default_name = f"{safe_name}_payment_history_{month_slug}.xlsx"
        file_path = filedialog.asksaveasfilename(
            parent=self,
            title="Export Payment History",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel Workbook", "*.xlsx")],
        )
        if not file_path:
            return

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Summary"

        items = sorted(selected_month["items"].keys(), key=lambda value: value.lower())
        if not items:
            items = ["Payment"]

        last_col = len(items) + 1
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        sheet["A1"] = self.client_name
        sheet["A2"] = "Date"
        for idx, item in enumerate(items, start=2):
            sheet.cell(row=2, column=idx, value=item)

        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        odd_fill = PatternFill(fill_type="solid", fgColor="EAF2FF")
        even_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
        header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
        summary_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
        center = Alignment(horizontal="center", vertical="center")

        title_font = Font(name="Segoe UI", bold=True, size=12)
        header_font = Font(name="Segoe UI", bold=True, size=10)
        body_font = Font(name="Segoe UI", size=10)

        for col in range(1, last_col + 1):
            top_cell = sheet.cell(row=1, column=col)
            header_cell = sheet.cell(row=2, column=col)
            top_cell.alignment = center
            top_cell.border = border
            top_cell.fill = header_fill
            header_cell.alignment = center
            header_cell.border = border
            header_cell.fill = header_fill
            header_cell.font = header_font
        sheet["A1"].font = title_font
        sheet["A2"].font = header_font

        current_row = 3
        days_in_month = calendar.monthrange(selected_month["year"], selected_month["month"])[1]
        for day in range(1, days_in_month + 1):
            fill = odd_fill if day % 2 == 1 else even_fill
            date_text = f"{datetime(selected_month['year'], selected_month['month'], day).strftime('%B')} {day}, {selected_month['year']}"
            sheet.cell(row=current_row, column=1, value=date_text)
            for col_index, item in enumerate(items, start=2):
                item_data = selected_month["items"].get(item, {"daily": {}})
                amount = float(item_data.get("daily", {}).get(day, 0.0) or 0.0)
                if amount > 0:
                    sheet.cell(row=current_row, column=col_index, value=amount)
                    sheet.cell(row=current_row, column=col_index).number_format = '#,##0.00'
            for col in range(1, last_col + 1):
                cell = sheet.cell(row=current_row, column=col)
                cell.border = border
                cell.alignment = center
                cell.fill = fill
                cell.font = body_font
            current_row += 1

        sheet.cell(row=current_row, column=1, value="Item Price")
        for col_index, item in enumerate(items, start=2):
            item_data = selected_month["items"].get(item, {})
            sheet.cell(row=current_row, column=col_index, value=float(item_data.get("item_price", 0.0)))
            sheet.cell(row=current_row, column=col_index).number_format = '#,##0.00'
        for col in range(1, last_col + 1):
            cell = sheet.cell(row=current_row, column=col)
            cell.border = border
            cell.alignment = center
            cell.fill = summary_fill
            cell.font = header_font
        current_row += 1

        sheet.cell(row=current_row, column=1, value="Balance")
        for col_index, item in enumerate(items, start=2):
            item_data = selected_month["items"].get(item, {})
            sheet.cell(row=current_row, column=col_index, value=float(item_data.get("remaining_balance", 0.0)))
            sheet.cell(row=current_row, column=col_index).number_format = '#,##0.00'
        for col in range(1, last_col + 1):
            cell = sheet.cell(row=current_row, column=col)
            cell.border = border
            cell.alignment = center
            cell.fill = summary_fill
            cell.font = header_font
        sheet.column_dimensions["A"].width = 18
        for col in range(2, last_col + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 14

        try:
            workbook.save(file_path)
        except Exception as exc:
            messagebox.showerror("Export Error", f"Could not export file:\n{exc}", parent=self)
            return

        messagebox.showinfo("Exported", f"Payment history exported to:\n{file_path}", parent=self)

    def _handle_close(self) -> None:
        if self._closed:
            return
        self._closed = True
        try:
            self.grab_release()
        except tk.TclError:
            pass
        try:
            if self.parent is not None and self.parent.winfo_exists():
                self.parent.grab_set()
                self.parent.focus_set()
        except tk.TclError:
            pass
        self.destroy()

    def _render_selected_month(self) -> None:
        for row_id in self.table.get_children():
            self.table.delete(row_id)

        selected_month = self._selected_month_entry()
        if selected_month is None:
            self.table.configure(columns=("date", "payment"))
            self.table.heading("date", text="Date", anchor="center")
            self.table.heading("payment", text="Payment", anchor="center")
            self.table.column("date", width=280, minwidth=180, anchor="center", stretch=True)
            self.table.column("payment", width=180, minwidth=140, anchor="center", stretch=True)
            self.table.insert("", "end", values=("No payment history found.", ""))
            return

        items = sorted(selected_month["items"].keys(), key=lambda value: value.lower())
        if not items:
            items = ["Payment"]

        column_ids = ["date"]
        self._column_to_item.clear()
        for idx, item in enumerate(items, start=1):
            col_id = f"item_{idx}"
            column_ids.append(col_id)
            self._column_to_item[col_id] = item

        self.table.configure(columns=tuple(column_ids))
        self.table.heading("date", text="Date", anchor="center")
        self.table.column("date", width=220, minwidth=180, anchor="center", stretch=True)
        for col_id, item in self._column_to_item.items():
            self.table.heading(col_id, text=item, anchor="center")
            self.table.column(col_id, width=140, minwidth=120, anchor="center", stretch=True)

        days_in_month = calendar.monthrange(selected_month["year"], selected_month["month"])[1]
        for day in range(1, days_in_month + 1):
            date_text = f"{datetime(selected_month['year'], selected_month['month'], day).strftime('%B')} {day}, {selected_month['year']}"
            row_values: list[str] = [date_text]
            for item in items:
                item_data = selected_month["items"].get(item, {"daily": {}})
                amount = float(item_data.get("daily", {}).get(day, 0.0) or 0.0)
                row_values.append(f"{amount:.2f}" if amount > 0 else "")
            self.table.insert(
                "",
                "end",
                values=tuple(row_values),
                tags=("odd",) if day % 2 == 1 else ("even",),
            )

        item_price_row = ["Item Price"]
        balance_row = ["Balance"]
        for item in items:
            item_data = selected_month["items"].get(item, {})
            item_price_row.append(f"{float(item_data.get('item_price', 0.0) or 0.0):.2f}")
            balance_row.append(f"{float(item_data.get('remaining_balance', 0.0) or 0.0):.2f}")
        self.table.insert("", "end", values=tuple(item_price_row), tags=("summary",))
        self.table.insert("", "end", values=tuple(balance_row), tags=("summary",))

    def _selected_month_entry(self) -> dict | None:
        selected_label = self.month_var.get().strip()
        for entry in self.month_entries:
            if entry["label"] == selected_label:
                return entry
        if self.month_entries:
            return self.month_entries[-1]
        return None

    def _build_month_data(self) -> list[dict]:
        now = datetime.now()
        month_map: dict[str, dict] = {}

        for group in self.grouped_history:
            event = group.get("event", {})
            payments = list(group.get("payments", []))

            event_date = self._parse_date(event.get("created_at"))
            if event_date is None:
                event_date = next(
                    (self._parse_date(payment.get("created_at")) for payment in payments if self._parse_date(payment.get("created_at")) is not None),
                    now,
                )
            event_month_key = event_date.strftime("%Y-%m")
            event_month_entry = month_map.setdefault(
                event_month_key,
                {
                    "key": event_month_key,
                    "label": event_date.strftime("%B %Y"),
                    "year": event_date.year,
                    "month": event_date.month,
                    "items": {},
                },
            )

            event_item = self._format_item_label(event.get("item"), "Payment")
            event_item_entry = event_month_entry["items"].setdefault(
                event_item,
                {"item_price": 0.0, "paid": 0.0, "daily": {}},
            )
            event_item_entry["item_price"] += float(event.get("amount", 0.0) or 0.0)

            for payment in payments:
                payment_date = self._parse_date(payment.get("created_at")) or event_date
                month_key = payment_date.strftime("%Y-%m")
                month_entry = month_map.setdefault(
                    month_key,
                    {
                        "key": month_key,
                        "label": payment_date.strftime("%B %Y"),
                        "year": payment_date.year,
                        "month": payment_date.month,
                        "items": {},
                    },
                )

                payment_item = self._format_item_label(payment.get("item"), event_item)
                payment_item_entry = month_entry["items"].setdefault(
                    payment_item,
                    {"item_price": 0.0, "paid": 0.0, "daily": {}},
                )
                payment_amount = float(payment.get("amount", 0.0) or 0.0)
                payment_item_entry["paid"] += payment_amount
                payment_item_entry["daily"][payment_date.day] = (
                    float(payment_item_entry["daily"].get(payment_date.day, 0.0) or 0.0) + payment_amount
                )

        entries = sorted(month_map.values(), key=lambda value: value["key"])
        for month_entry in entries:
            for item_name, item_data in month_entry["items"].items():
                _ = item_name
                item_price = float(item_data.get("item_price", 0.0) or 0.0)
                paid = float(item_data.get("paid", 0.0) or 0.0)
                item_data["remaining_balance"] = max(0.0, item_price - paid)
        return entries

    @staticmethod
    def _format_item_label(item_value: str | None, fallback: str) -> str:
        label = str(item_value or "").strip()
        if label:
            return label
        return str(fallback or "").strip() or "Payment"

    @staticmethod
    def _parse_date(raw_value: str | None) -> datetime | None:
        raw = str(raw_value or "").strip()
        if not raw:
            return None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(raw, fmt)
            except ValueError:
                continue
        return None